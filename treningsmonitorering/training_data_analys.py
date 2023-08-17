# -*- coding: utf-8 -*-
"""
Created on Sat Apr 29 10:32:03 2023

@author: Roger
"""
import os
import django
os.environ.setdefault('DJANGO_SETTINGS_MODULE', 'treningsmonitorering.settings')
django.setup()
from db_interaction import update_rm
from trening_db.models import ResistanceExercise, DailyTrainingVolume, WeeklyTrainingVolum, TrainingRepMax, DailyVariables, \
    WeeklyTrainingVolumeWeight, DailyTrainingVolumeWeight
import calendar
from openpyxl import load_workbook
import datetime
from time import time
import re


Days = ("Monday", "Tuesday", "Wednesday", "Thursday", "Friday", "Saturday", "Sunday")
Daily_variables = ("Øvelse", "Kroppsvekt", "Skade", "Søvntid", "Søvnkvalitet", "Skadetype", "Dato", "^", None)
Variables = ("^",
             "Monday",
             "Tuesday",
             "Wednesday",
             "Thursday",
             "Friday",
             "Saturday",
             "Sunday",
             "Øvelse",
             "Kroppsvekt",
             "Skade",
             "Søvntid",
             "Søvnkvalitet",
             "Skadetype",
             "Dato",
             None)


def performance(func):
    def wrapper(*args, **kwargs):
        t1 = time()
        result = func(*args, **kwargs)
        t2 = time()
        time_used = t2-t1

        print(f'''function "{func.__name__}" took: {time_used} ms''')
        return result
    return wrapper

@performance
def get_distribution_all_days(sheet_index, file_path, name, age, gender):

    workbook = load_workbook(filename=file_path)

    def flexible_exercises(exercise):
        if type(exercise) == str:
            whitespace_gone_ex = exercise.strip()
            caps_letters_gone = whitespace_gone_ex.lower()

            return caps_letters_gone

    def find_exercises(sheet_index=None):

        sheet = workbook[sheet_index]
        list_of_exercises = []
        cells = [cell.value for cell in sheet['A'] if cell.value != None]
        dates_re = re.compile(r'^([0-9]{2})?[0-9]{2}(\/|-)(1[0-2]|0?[1-9])\2(3[01]|[12][0-9]|0?[1-9])$')
        comments = re.compile(r"\^+")

        for item in cells:
            flex_ex = flexible_exercises(item)
            comment = comments.search(item)
            date = dates_re.search(item)
            if comment or date or item in Daily_variables or item in Days or flex_ex in list_of_exercises:
                continue
            list_of_exercises.append(flex_ex)

        return list_of_exercises

    def find_one_rm(sheet_index=None):

        complex_exercises = re.compile(r"\b(?:[a-zA-ZÅØÆåøæ]+\s*\+\s*)+[a-zA-ZÅØÆåøæ]+\b")
        sheet = workbook['RM oversikt']
        one_RM_exercise = {}

        for exercise in find_exercises(sheet_index):
            complex_ex = complex_exercises.search(exercise)
            if complex_ex:
                exrss = split_complex_exercises(exercise)
                for exr in exrss:
                    for cell in sheet['A']:
                        flex_ex = flexible_exercises(cell.value)
                        if flex_ex == exr and flex_ex not in Daily_variables:
                            row_of_interest = cell.row
                            one_RM = sheet['B' + str(row_of_interest)].value
                            one_RM_exercise[exr] = one_RM
                continue

            for cell in sheet['A']:
                flex_ex = flexible_exercises(cell.value)
                if flex_ex == exercise and flex_ex not in Daily_variables:
                    row_of_interest = cell.row
                    one_RM = sheet['B' + str(row_of_interest)].value
                    one_RM_exercise[exercise] = one_RM

        return one_RM_exercise

    def find_days(sheet_index):

        sheet = workbook[sheet_index]
        training_days = {
            "Monday": [],
            "Tuesday": [],
            "Wednesday": [],
            "Thursday": [],
            "Friday": [],
            "Saturday": [],
            "Sunday": [],
        }

        for day in Days:
            for cell in sheet["A"]:
                if cell.value == day:
                    training_days[day].append(cell.row)

        return training_days

    def find_daily_variables(sheet_index, name):

        sheet = workbook[sheet_index]
        daily_variables = {}
        dayz = [day for day in find_days(sheet_index).keys()]
        ids = list(ResistanceExercise.objects.filter(student__name=name).values_list('id', flat=True))
        stud_id = ids[0]
        instance = ResistanceExercise.objects.get(pk=stud_id)

        for index, row_number in enumerate(find_days(sheet_index).values()):
            variables = [daily_vars.value for daily_vars in sheet[row_number[0] - 1] if daily_vars.value != None]
            daily_variables[dayz[index]] = (variables)
            vrs = DailyVariables(mode_of_exercise=instance,
                                 date=variables[0],
                                 weight=variables[1],
                                 injury_status=variables[2],
                                 injury_type=variables[3],
                                 sleep_quality=variables[4],
                                 sleep_duration=variables[5],
                                 training_period=variables[6],
                                 rpe=variables[7],
                                 prs=variables[8])
            vrs.save()

        return daily_variables

    def find_dates(sheet_index):

        sheet = workbook[sheet_index]
        daily_variables = {}
        dayz = [day for day in find_days(sheet_index).keys()]

        for index, row_number in enumerate(find_days(sheet_index).values()):
            variables = [daily_vars.value for daily_vars in sheet[row_number[0] - 1] if daily_vars.value != None]
            daily_variables[dayz[index]] = (variables)

        format = "%Y-%m-%d"
        dates = [datetime.datetime.strptime(date[1][0], format) for date in daily_variables.items()]

        return dates

    def find_daily_exercises(sheet_index):

        sheet = workbook[sheet_index]
        Col_of_interest = sheet["A"]
        daily_exercises = {
            "Monday": [],
            "Tuesday": [],
            "Wednesday": [],
            "Thursday": [],
            "Friday": [],
            "Saturday": [],
            "Sunday": [],
        }

        for days, row_num in find_days(sheet_index).items():
            for exercise in Col_of_interest[row_num[0]:row_num[1]]:
                if exercise.value not in Variables:
                    daily_exercises[days].append(exercise.value)

        return daily_exercises

    def split_complex_exercises(exercise=None):
        whitespace_exercises = re.compile(r"\s*(?:[a-zA-ZÅØÆåøæ\s]*)\b")
        x = whitespace_exercises.findall(exercise)
        x2 = [i for i in x if i]

        complex_reps_distribution = []

        for i, exrs in enumerate(x2):
            trimmed_exercise = exrs.strip()

            complex_reps_distribution.append(trimmed_exercise)

        return complex_reps_distribution

    def split_complex_reps(reps=None):
        find_complex_reps = re.compile(r"(?:\b\d\s*\+\s*)+\d\b")
        find_numbers = re.compile(r"\d")

        complex_reps = []

        com_reps = find_complex_reps.search(reps)
        if com_reps:
            complex_reps.append(reps)
        nrs = []
        for x in complex_reps:
            number = find_numbers.findall(x)
            for num in number:
                nr = int(num)
                nrs.append(nr)

        return nrs

    def find_daily_weight(sheet_index):
        complex_exercises = re.compile(r"\b(?:[a-zA-ZÅØÆåøæ]+\s*\+\s*)+[a-zA-ZÅØÆåøæ]+\b")
        comments = re.compile(r"\^+")
        sheet = workbook[sheet_index]
        Col_of_interest = sheet["A"]
        cmnts = [cell.value for cell in Col_of_interest if cell.value != None and comments.search(cell.value)]

        daily_training_rundown_weight = {}

        for day, row_number in find_days(sheet_index).items():
            for exercise in Col_of_interest[row_number[0]:row_number[1]]:
                if exercise.value not in Variables and exercise.value not in cmnts:
                    exercise_row_location = exercise.row
                    complex_ex = complex_exercises.search(exercise.value)
                    if complex_ex:
                        exercises = split_complex_exercises(exercise.value)
                        for exr1 in exercises:
                            exr = flexible_exercises(exr1)
                            for weight in sheet[exercise_row_location]:
                                if weight.value not in Variables and weight.value != exercise.value and (
                                        day, exr) not in daily_training_rundown_weight.keys():
                                    daily_training_rundown_weight[day, exr] = [weight.value]
                                elif weight.value not in Variables and weight.value != exercise.value and (
                                        day, exr) in daily_training_rundown_weight.keys():
                                    daily_training_rundown_weight[(day, exr)].append(weight.value)

                        continue

                    exr = flexible_exercises(exercise.value)
                    for weight in sheet[exercise_row_location]:
                        if weight.value not in Variables and weight.value != exercise.value and (
                                day, exr) not in daily_training_rundown_weight.keys() \
                                and weight.value not in cmnts:
                            daily_training_rundown_weight[day, exr] = [weight.value]
                        elif weight.value not in Variables and weight.value != exercise.value and (
                                day, exr) in daily_training_rundown_weight.keys() \
                            and weight.value not in cmnts:
                            daily_training_rundown_weight[(day, exr)].append(weight.value)

        return daily_training_rundown_weight

    def find_daily_reps(sheet_index):
        complex_exercises = re.compile(r"\b(?:[a-zA-ZÅØÆåøæ]+\s*\+\s*)+[a-zA-ZÅØÆåøæ]+\b")
        comments = re.compile(r"\^+")
        sheet = workbook[sheet_index]
        Col_of_interest = sheet["A"]
        cmnts = [cell.value for cell in Col_of_interest if cell.value != None and comments.search(cell.value)]

        daily_training_rundown_reps = {}

        for day, row_number in find_days(sheet_index).items():
            for exercise in Col_of_interest[row_number[0]:row_number[1]]:
                if exercise.value not in Variables and exercise.value not in cmnts:
                    exercise_row_location = exercise.row
                    complex_ex = complex_exercises.search(exercise.value)
                    if complex_ex:
                        exercises = split_complex_exercises(exercise.value)
                        for i, exr1 in enumerate(exercises):
                            exr = flexible_exercises(exr1)
                            for reps in sheet[exercise_row_location + 1]:
                                if reps.value not in Variables and reps.value != exercise.value and (
                                        day, exr) not in daily_training_rundown_reps.keys():
                                    if type(reps.value) == str:
                                        comment = comments.search(reps.value)
                                        if comment:
                                            continue
                                        rp = split_complex_reps(reps.value)
                                        daily_training_rundown_reps[day, exr] = [rp[i]]
                                elif reps.value not in Variables and reps.value != exercise.value and (
                                        day, exr) in daily_training_rundown_reps.keys():
                                    if type(reps.value) == str:
                                        rp = split_complex_reps(reps.value)
                                        daily_training_rundown_reps[(day, exr)].append(rp[i])

                        continue
                    exr = flexible_exercises(exercise.value)
                    for reps in sheet[exercise_row_location + 1]:
                        if reps.value not in Variables and reps.value != exercise.value and (
                                day, exr) not in daily_training_rundown_reps.keys() \
                                and reps.value not in cmnts:
                            daily_training_rundown_reps[day, exr] = [reps.value]
                        elif reps.value not in Variables and reps.value != exercise.value and (
                                day, exr) in daily_training_rundown_reps.keys() and reps.value not in cmnts:
                            daily_training_rundown_reps[(day, exr)].append(reps.value)

        return daily_training_rundown_reps

    # @performance
    def find_daily_training_distribution(day, sheet_index, date):
        one_rm_values = find_one_rm(sheet_index)
        exercise_weight = {}
        exercise_reps = {}
        total_daily_distribution_reps = {"0-10%": 0,
                                        "11-20%": 0,
                                        "21-30%": 0,
                                        "31-40%": 0,
                                        "41-50%": 0,
                                        "51-60%": 0,
                                        "61-70%": 0,
                                        "71-80%": 0,
                                        "81-90%": 0,
                                        "91-100%": 0,
                                        "101-110%": 0,
                                        "111-120%": 0,
                                        "121-130%": 0
                                        }
        total_daily_distribution_weight = {"0-10%": 0,
                                          "11-20%": 0,
                                          "21-30%": 0,
                                          "31-40%": 0,
                                          "41-50%": 0,
                                          "51-60%": 0,
                                          "61-70%": 0,
                                          "71-80%": 0,
                                          "81-90%": 0,
                                          "91-100%": 0,
                                          "101-110%": 0,
                                          "111-120%": 0,
                                          "121-130%": 0
                                          }
        sum_of_total_daily_reps = 0
        sum_of_total_daily_weight = 0
        intensity_zones = {
            "0-10%": [0.0, 0.1],
            "11-20%": [0.11, 0.2],
            "21-30%": [0.21, 0.3],
            "31-40%": [0.31, 0.4],
            "41-50%": [0.41, 0.5],
            "51-60%": [0.51, 0.6],
            "61-70%": [0.61, 0.7],
            "71-80%": [0.71, 0.8],
            "81-90%": [0.81, 0.9],
            "91-100%": [0.91, 1.0],
            "101-110%": [1.01, 1.1],
            "111-120%": [1.11, 1.2],
            "121-130%": [1.2, 1.3]
        }

        def find_variables(variable):

            if variable == find_daily_weight(sheet_index).items():
                for exercise_volume in variable:
                    if day in exercise_volume[0]:
                        exercise_weight[exercise_volume[0][1]] = exercise_volume[1]

                return exercise_weight

            elif variable == find_daily_reps(sheet_index).items():
                for exercise_volume in variable:
                    if day in exercise_volume[0]:
                        exercise_reps[exercise_volume[0][1]] = exercise_volume[1]

                return exercise_reps

        find_variables(find_daily_weight(sheet_index).items())
        find_variables(find_daily_reps(sheet_index).items())

        student_obj_ids = list(ResistanceExercise.objects.filter(student__name=name).values_list('id', flat=True))
        stud_obj = student_obj_ids[0]
        instance = ResistanceExercise.objects.get(pk=stud_obj)

        for k, exercise in enumerate(exercise_weight.keys()):
            exercise_volume_distribution_reps = {"0-10%": 0,
                                                "11-20%": 0,
                                                "21-30%": 0,
                                                "31-40%": 0,
                                                "41-50%": 0,
                                                "51-60%": 0,
                                                "61-70%": 0,
                                                "71-80%": 0,
                                                "81-90%": 0,
                                                "91-100%": 0,
                                                "101-110%": 0,
                                                "111-120%": 0,
                                                "121-130%": 0
                                            }
            exercise_volume_distribution_weight = {"0-10%": 0,
                                                  "11-20%": 0,
                                                  "21-30%": 0,
                                                  "31-40%": 0,
                                                  "41-50%": 0,
                                                  "51-60%": 0,
                                                  "61-70%": 0,
                                                  "71-80%": 0,
                                                  "81-90%": 0,
                                                  "91-100%": 0,
                                                  "101-110%": 0,
                                                  "111-120%": 0,
                                                  "121-130%": 0
                                                  }
            total_exercise_reps = 0
            total_exercise_weight = 0

            for index, weight in enumerate(exercise_weight[exercise]):
                for interval in intensity_zones.items():
                    if exercise in one_rm_values.keys() and weight in range(
                            (int(one_rm_values[exercise] * interval[1][0])),
                            (int(one_rm_values[exercise] * interval[1][1]))):
                        exercise_volume_distribution_reps[interval[0]] += (exercise_reps[exercise][index])
                        exercise_volume_distribution_weight[interval[0]] += weight * (exercise_reps[exercise][index])
                        total_daily_distribution_reps[interval[0]] += (exercise_reps[exercise][index])
                        total_daily_distribution_weight[interval[0]] += weight * (exercise_reps[exercise][index])

                        r3ps = exercise_reps[exercise][index]  # see if it is possible to move this up to the line under
                        # enumerate(exercise_weight[exercise]):
                        update_rm(name, exercise, date, weight, r3ps)
                        continue

                total_exercise_reps += exercise_reps[exercise][index]
                total_exercise_weight += weight * exercise_reps[exercise][index]
            sum_of_total_daily_reps += total_exercise_reps  # OBS This also adds the values that are beyond the scope of investigation (higher than 130% of 1RM), even tho the values are not added into the intensity distribution
            sum_of_total_daily_weight += total_exercise_weight


            volume = DailyTrainingVolume(mode_of_exercise=instance,
                                         date=date,
                                         day=day,
                                         exercise=exercise,
                                         total_exercise_reps_volume=total_exercise_reps,
                                         intensity_zone_1=exercise_volume_distribution_reps["0-10%"],
                                         intensity_zone_2=exercise_volume_distribution_reps["11-20%"],
                                         intensity_zone_3=exercise_volume_distribution_reps["21-30%"],
                                         intensity_zone_4=exercise_volume_distribution_reps["31-40%"],
                                         intensity_zone_5=exercise_volume_distribution_reps["41-50%"],
                                         intensity_zone_6=exercise_volume_distribution_reps["51-60%"],
                                         intensity_zone_7=exercise_volume_distribution_reps["61-70%"],
                                         intensity_zone_8=exercise_volume_distribution_reps["71-80%"],
                                         intensity_zone_9=exercise_volume_distribution_reps["81-90%"],
                                         intensity_zone_10=exercise_volume_distribution_reps["91-100%"],
                                         intensity_zone_11=exercise_volume_distribution_reps["101-110%"],
                                         intensity_zone_12=exercise_volume_distribution_reps["111-120%"],
                                         intensity_zone_13=exercise_volume_distribution_reps["121-130%"])

            volume4 = DailyTrainingVolumeWeight(mode_of_exercise=instance,
                                                date=date,
                                                day=day,
                                                exercise=exercise,
                                                total_exercise_weight_volume=total_exercise_weight,
                                                intensity_zone_1=exercise_volume_distribution_weight["0-10%"],
                                                intensity_zone_2=exercise_volume_distribution_weight["11-20%"],
                                                intensity_zone_3=exercise_volume_distribution_weight["21-30%"],
                                                intensity_zone_4=exercise_volume_distribution_weight["31-40%"],
                                                intensity_zone_5=exercise_volume_distribution_weight["41-50%"],
                                                intensity_zone_6=exercise_volume_distribution_weight["51-60%"],
                                                intensity_zone_7=exercise_volume_distribution_weight["61-70%"],
                                                intensity_zone_8=exercise_volume_distribution_weight["71-80%"],
                                                intensity_zone_9=exercise_volume_distribution_weight["81-90%"],
                                                intensity_zone_10=exercise_volume_distribution_weight["91-100%"],
                                                intensity_zone_11=exercise_volume_distribution_weight["101-110%"],
                                                intensity_zone_12=exercise_volume_distribution_weight["111-120%"],
                                                intensity_zone_13=exercise_volume_distribution_weight["121-130%"])

            volume.save()
            volume4.save()

        volume2 = DailyTrainingVolume(mode_of_exercise=instance,
                                      date=date,
                                      day=day,
                                      exercise='Total Reps',
                                      total_exercise_reps_volume= sum_of_total_daily_reps,
                                      intensity_zone_1=total_daily_distribution_reps["0-10%"],
                                      intensity_zone_2=total_daily_distribution_reps["11-20%"],
                                      intensity_zone_3=total_daily_distribution_reps["21-30%"],
                                      intensity_zone_4=total_daily_distribution_reps["31-40%"],
                                      intensity_zone_5=total_daily_distribution_reps["41-50%"],
                                      intensity_zone_6=total_daily_distribution_reps["51-60%"],
                                      intensity_zone_7=total_daily_distribution_reps["61-70%"],
                                      intensity_zone_8=total_daily_distribution_reps["71-80%"],
                                      intensity_zone_9=total_daily_distribution_reps["81-90%"],
                                      intensity_zone_10=total_daily_distribution_reps["91-100%"],
                                      intensity_zone_11=total_daily_distribution_reps["101-110%"],
                                      intensity_zone_12=total_daily_distribution_reps["111-120%"],
                                      intensity_zone_13=total_daily_distribution_reps["121-130%"])

        volume3 = DailyTrainingVolumeWeight(mode_of_exercise=instance,
                                            date=date,
                                            day=day,
                                            exercise='Total Weight',
                                            total_exercise_weight_volume=sum_of_total_daily_weight,
                                            intensity_zone_1=total_daily_distribution_weight["0-10%"],
                                            intensity_zone_2=total_daily_distribution_weight["11-20%"],
                                            intensity_zone_3=total_daily_distribution_weight["21-30%"],
                                            intensity_zone_4=total_daily_distribution_weight["31-40%"],
                                            intensity_zone_5=total_daily_distribution_weight["41-50%"],
                                            intensity_zone_6=total_daily_distribution_weight["51-60%"],
                                            intensity_zone_7=total_daily_distribution_weight["61-70%"],
                                            intensity_zone_8=total_daily_distribution_weight["71-80%"],
                                            intensity_zone_9=total_daily_distribution_weight["81-90%"],
                                            intensity_zone_10=total_daily_distribution_weight["91-100%"],
                                            intensity_zone_11=total_daily_distribution_weight["101-110%"],
                                            intensity_zone_12=total_daily_distribution_weight["111-120%"],
                                            intensity_zone_13=total_daily_distribution_weight["121-130%"])

        volume2.save()
        volume3.save()

        return total_daily_distribution_reps

    dates = find_dates(sheet_index)
    dts = set([dt.pr_date for dt in TrainingRepMax.objects.filter(mode_of_exercise__student__name=name)])

    for day in Days:
        for dt in dates:
            wkday = calendar.day_name[dt.weekday()]
            date = dt.date()
            if day == wkday and date not in dts:
                find_daily_training_distribution(day, sheet_index, date)

@performance
def get_distribution_all_weeks(sheet_index, file_path, name, age, gender):

    workbook = load_workbook(filename=file_path)

    def flexible_exercises(exercise):
        if type(exercise) == str:
            whitespace_gone_ex = exercise.strip()
            caps_letters_gone = whitespace_gone_ex.lower()

            return caps_letters_gone

    def find_exercises(sheet_index=None):

        sheet = workbook[sheet_index]
        list_of_exercises = []
        cells = [cell.value for cell in sheet['A'] if cell.value != None]
        dates_re = re.compile(r'^([0-9]{2})?[0-9]{2}(\/|-)(1[0-2]|0?[1-9])\2(3[01]|[12][0-9]|0?[1-9])$')
        comments = re.compile(r"\^+")

        for item in cells:
            flex_ex = flexible_exercises(item)
            comment = comments.search(item)
            date = dates_re.search(item)
            if comment or date or item in Daily_variables or item in Days or flex_ex in list_of_exercises:
                continue
            list_of_exercises.append(flex_ex)

        return list_of_exercises

    def find_weekly_exercise_locations(sheet_index):

        sheet = workbook[sheet_index]
        exercise_location = {}
        exercises = find_exercises(sheet_index)
        comments = re.compile(r"\^+")
        cmnts = [cell.value for cell in sheet["A"] if cell.value != None and comments.search(cell.value)]

        for exercise in exercises:
            for cell in sheet['A']:
                exr = flexible_exercises(cell.value)
                if cell.value not in cmnts and exr == exercise and exercise not in exercise_location:
                    exercise_location[exercise] = [cell.row]
                elif cell.value not in cmnts and exr == exercise and exercise in exercise_location:
                    exercise_location[exercise].append(cell.row)

        return exercise_location

    def split_complex_exercises(exercise=None):
        whitespace_exercises = re.compile(r"\s*(?:[a-zA-ZÅØÆåøæ\s]*)\b")
        x = whitespace_exercises.findall(exercise)
        x2 = [i for i in x if i]

        complex_reps_distribution = []

        for i, exrs in enumerate(x2):
            trimmed_exercise = exrs.strip()

            complex_reps_distribution.append(trimmed_exercise)

        return complex_reps_distribution

    def split_complex_reps(reps=None):
        regex_finds_reps = re.compile(r"(?:\b\d\s*\+\s*)+\d\b")
        find_numbers = re.compile(r"\d")

        complex_reps = []

        com_reps = regex_finds_reps.search(reps)
        if com_reps:
            complex_reps.append(reps)
        nrs = []
        for x in complex_reps:
            number = find_numbers.findall(x)
            for num in number:
                nr = int(num)
                nrs.append(nr)

        return nrs

    def find_weekly_weight_pr_set(sheet_index):
        complex_exercises = re.compile(r"\b(?:[a-zA-ZÅØÆåøæ]+\s*\+\s*)+[a-zA-ZÅØÆåøæ]+\b")
        comments = re.compile(r"\^+")
        sheet = workbook[sheet_index]
        weight_pr_exercise = {}
        exercises = find_exercises(sheet_index)
        cmnts = [cell.value for cell in sheet["A"] if cell.value != None and comments.search(cell.value)]
        sheet_A_values = [cell.value for cell in sheet['A'] if cell.value not in cmnts]

        for exercise in exercises:
            for row_number in find_weekly_exercise_locations(sheet_index)[exercise]:
                complex_ex = complex_exercises.search(exercise)
                if complex_ex:
                    exrss = split_complex_exercises(exercise)
                    for i, exr1 in enumerate(exrss):
                        exr = flexible_exercises(exr1)
                        for weight in sheet[row_number]:
                            if weight.value not in exercises and weight.value != None and exr not in weight_pr_exercise\
                                    and weight.value not in cmnts and weight.value not in sheet_A_values:
                                weight_pr_exercise[exr] = [weight.value]
                            elif weight.value not in exercises and weight.value != None and exr in weight_pr_exercise \
                                    and weight.value not in cmnts and weight.value not in sheet_A_values:
                                weight_pr_exercise[exr].append(weight.value)
                    continue

                exr = flexible_exercises(exercise)
                for cell in sheet[row_number]:
                    if cell.value not in exercises and cell.value != None and exr not in weight_pr_exercise and \
                            cell.value not in cmnts and cell.value not in sheet_A_values:
                        weight_pr_exercise[exr] = [cell.value]
                    elif cell.value not in exercises and cell.value != None and exr in weight_pr_exercise \
                            and cell.value not in cmnts  and cell.value not in sheet_A_values:
                        weight_pr_exercise[exr].append(cell.value)

        return weight_pr_exercise

    def find_weekly_reps_pr_set(sheet_index):
        complex_exercises = re.compile(r"\b(?:[a-zA-ZÅØÆåøæ]+\s*\+\s*)+[a-zA-ZÅØÆåøæ]+\b")
        comments = re.compile(r"\^+")
        sheet = workbook[sheet_index]
        reps_pr_set = {}
        exercises = find_exercises(sheet_index)
        cmnts = [cell.value for cell in sheet["A"] if cell.value != None and comments.search(cell.value)]

        for exercise in exercises:
            for row_number in find_weekly_exercise_locations(sheet_index)[exercise]:
                complex_ex = complex_exercises.search(exercise)
                if complex_ex:
                    exrss = split_complex_exercises(exercise)
                    for i, exr1 in enumerate(exrss):
                        exr = flexible_exercises(exr1)
                        for reps in sheet[row_number + 1]:
                            if reps.value != None and reps.value not in exercises and exr not in reps_pr_set and \
                                    reps.value not in cmnts:
                                if type(reps.value) == str:  # and int(reps.value[0]) == int:
                                    rp = split_complex_reps(reps.value)
                                    reps_pr_set[exr] = [rp[i]]

                            elif reps.value != None and reps.value not in exercises and exr in reps_pr_set:
                                comment = comments.search(reps.value)
                                if comment:
                                    continue
                                if type(reps.value) == str:
                                    rp = split_complex_reps(reps.value)
                                    reps_pr_set[exr].append(rp[i])
                    continue

                exr = flexible_exercises(exercise)
                for cell in sheet[row_number + 1]:
                    if cell.value != None and cell.value not in exercises and exr not in reps_pr_set and \
                            cell.value not in cmnts:
                        reps_pr_set[exr] = [cell.value]
                    elif cell.value != None and cell.value not in exercises and exr in reps_pr_set and \
                        cell.value not in cmnts:
                        reps_pr_set[exr].append(cell.value)

        return reps_pr_set

    def find_days(sheet_index):

        sheet = workbook[sheet_index]
        training_days = {
            "Monday": [],
            "Tuesday": [],
            "Wednesday": [],
            "Thursday": [],
            "Friday": [],
            "Saturday": [],
            "Sunday": [],
        }

        for day in Days:
            for cell in sheet["A"]:
                if cell.value == day:
                    training_days[day].append(cell.row)

        return training_days

    def find_dates(sheet_index):

        sheet = workbook[sheet_index]
        daily_variables = {}
        dayz = [day for day in find_days(sheet_index).keys()]

        for index, row_number in enumerate(find_days(sheet_index).values()):
            variables = [daily_vars.value for daily_vars in sheet[row_number[0] - 1] if daily_vars.value != None]
            daily_variables[dayz[index]] = (variables)


        format = "%Y-%m-%d"
        dates = [datetime.datetime.strptime(date[1][0], format) for date in daily_variables.items()]

        return dates

    def find_one_rm(sheet_index=None):

        complex_exercises = re.compile(r"\b(?:[a-zA-ZÅØÆåøæ]+\s*\+\s*)+[a-zA-ZÅØÆåøæ]+\b")
        sheet = workbook['RM oversikt']
        one_RM_exercise = {}

        for exercise in find_exercises(sheet_index):
            complex_ex = complex_exercises.search(exercise)
            if complex_ex:
                exrss = split_complex_exercises(exercise)
                for exr in exrss:
                    for cell in sheet['A']:
                        flex_ex = flexible_exercises(cell.value)
                        if flex_ex == exr and flex_ex not in Daily_variables:
                            row_of_interest = cell.row
                            one_RM = sheet['B' + str(row_of_interest)].value
                            one_RM_exercise[exr] = one_RM
                continue

            for cell in sheet['A']:
                flex_ex = flexible_exercises(cell.value)
                if flex_ex == exercise and flex_ex not in Daily_variables:
                    row_of_interest = cell.row
                    one_RM = sheet['B' + str(row_of_interest)].value
                    one_RM_exercise[exercise] = one_RM

        return one_RM_exercise

    # @performance
    def find_weekly_training_distribution(sheet_index):  # creates objects containing the exercise name and the training distribution

        sht_name = re.compile('\d')
        try:
            week = sht_name.findall(sheet_index)
            if len(week) < 2:
                week = sht_name.findall(sheet_index)[0]
            else:
                weeks = sht_name.findall(sheet_index)
                week = weeks[0] + weeks[1]

        except IndexError:
            week = 1

        complex_exercises = re.compile(r"\b(?:[a-zA-ZÅØÆåøæ]+\s*\+\s*)+[a-zA-ZÅØÆåøæ]+\b")
        exercises = find_exercises(sheet_index)
        weekly_weight = find_weekly_weight_pr_set(sheet_index)
        one_RM = find_one_rm(sheet_index)
        wekly_reps = find_weekly_reps_pr_set(sheet_index)
        intervals = {
            "0-10%": (0.0, 0.10),
            "11-20%": (0.11, 0.20),
            "21-30%": (0.21, 0.30),
            "31-40%": (0.31, 0.4),
            "41-50%": (0.41, 0.5),
            "51-60%": (0.51, 0.6),
            "61-70%": (0.61, 0.7),
            "71-80%": (0.71, 0.8),
            "81-90%": (0.81, 0.9),
            "91-100%": (0.91, 1.0),
            "101-110%": (1.01, 1.1),
            "111-120%": (1.11, 1.2),
            "121-130%": (1.21, 1.3)
        }

        total_training_distribution_reps = {
            "Total reps": 0,
            "0-10%": 0,
            "11-20%": 0,
            "21-30%": 0,
            "31-40%": 0,
            "41-50%": 0,
            "51-60%": 0,
            "61-70%": 0,
            "71-80%": 0,
            "81-90%": 0,
            "91-100%": 0,
            "101-110%": 0,
            "111-120%": 0,
            "121-130%": 0
        }
        total_training_distribution_weight = {
            "Total weight": 0,
            "0-10%": 0,
            "11-20%": 0,
            "21-30%": 0,
            "31-40%": 0,
            "41-50%": 0,
            "51-60%": 0,
            "61-70%": 0,
            "71-80%": 0,
            "81-90%": 0,
            "91-100%": 0,
            "101-110%": 0,
            "111-120%": 0,
            "121-130%": 0
        }

        student_obj_ids = list(ResistanceExercise.objects.filter(student__name=name).values_list('id', flat=True))
        stud_obj = student_obj_ids[0]
        instance = ResistanceExercise.objects.get(pk=stud_obj)

        for exercise in exercises:
            training_distribution_reps = {
                "Total reps": 0,
                "0-10%": 0,
                "11-20%": 0,
                "21-30%": 0,
                "31-40%": 0,
                "41-50%": 0,
                "51-60%": 0,
                "61-70%": 0,
                "71-80%": 0,
                "81-90%": 0,
                "91-100%": 0,
                "101-110%": 0,
                "111-120%": 0,
                "121-130%": 0
            }
            training_distribution_weight = {
                'Total weight': 0,
                "0-10%": 0,
                "11-20%": 0,
                "21-30%": 0,
                "31-40%": 0,
                "41-50%": 0,
                "51-60%": 0,
                "61-70%": 0,
                "71-80%": 0,
                "81-90%": 0,
                "91-100%": 0,
                "101-110%": 0,
                "111-120%": 0,
                "121-130%": 0
            }
            complex_ex = complex_exercises.search(exercise)
            if complex_ex:
                exrss = split_complex_exercises(exercise)
                for exr in exrss:
                    complex_training_distribution_reps = {
                        "Total reps": 0,
                        "0-10%": 0,
                        "11-20%": 0,
                        "21-30%": 0,
                        "31-40%": 0,
                        "41-50%": 0,
                        "51-60%": 0,
                        "61-70%": 0,
                        "71-80%": 0,
                        "81-90%": 0,
                        "91-100%": 0,
                        "101-110%": 0,
                        "111-120%": 0,
                        "121-130%": 0
                    }
                    complex_training_distribution_weight = {
                        "Total reps": 0,
                        'Total weight': 0,
                        "0-10%": 0,
                        "11-20%": 0,
                        "21-30%": 0,
                        "31-40%": 0,
                        "41-50%": 0,
                        "51-60%": 0,
                        "61-70%": 0,
                        "71-80%": 0,
                        "81-90%": 0,
                        "91-100%": 0,
                        "101-110%": 0,
                        "111-120%": 0,
                        "121-130%": 0
                    }
                    for key_weights in weekly_weight.keys():
                        for key_one_rm in one_RM.keys():
                            if exr == key_weights and exr == key_one_rm:
                                for index, weight in enumerate(weekly_weight[exr]):
                                    for interval in intervals.items():

                                        if weight in range((int(one_RM[key_one_rm] * interval[1][0])),
                                                           (int(one_RM[key_one_rm] * interval[1][1]))):
                                            complex_training_distribution_reps[interval[0]] += (wekly_reps[exr][index])
                                            complex_training_distribution_weight[interval[0]] += weight * (wekly_reps[exr][index])
                                            complex_training_distribution_weight['Total weight'] += weight * wekly_reps[exr][index]
                                            complex_training_distribution_reps["Total reps"] += (wekly_reps[exr][index])
                                            total_training_distribution_reps[interval[0]] += (wekly_reps[exr][index])
                                            total_training_distribution_reps["Total reps"] += (wekly_reps[exr][index])
                                            total_training_distribution_weight['Total weight'] += weight * wekly_reps[exr][index]
                                            total_training_distribution_weight[interval[0]] += weight * wekly_reps[exr][index]
                                            continue

                    volume3 = WeeklyTrainingVolum(mode_of_exercise=instance,
                                                  wknr=wknr[0],
                                                  exercise=exr,
                                                  total_intensity_zone=complex_training_distribution_reps["Total reps"],
                                                  intensity_zone_1=complex_training_distribution_reps["0-10%"],
                                                  intensity_zone_2=complex_training_distribution_reps["11-20%"],
                                                  intensity_zone_3=complex_training_distribution_reps["21-30%"],
                                                  intensity_zone_4=complex_training_distribution_reps["31-40%"],
                                                  intensity_zone_5=complex_training_distribution_reps["41-50%"],
                                                  intensity_zone_6=complex_training_distribution_reps["51-60%"],
                                                  intensity_zone_7=complex_training_distribution_reps["61-70%"],
                                                  intensity_zone_8=complex_training_distribution_reps["71-80%"],
                                                  intensity_zone_9=complex_training_distribution_reps["81-90%"],
                                                  intensity_zone_10=complex_training_distribution_reps["91-100%"],
                                                  intensity_zone_11=complex_training_distribution_reps["101-110%"],
                                                  intensity_zone_12=complex_training_distribution_reps["111-120%"],
                                                  intensity_zone_13=complex_training_distribution_reps["121-130%"],
                                                  week=week)

                    volume4 = WeeklyTrainingVolumeWeight(mode_of_exercise=instance,
                                                         wknr=wknr[0],
                                                         exercise=exr,
                                                         total_exercise_weight_volume=complex_training_distribution_weight["Total weight"],
                                                         intensity_zone_1=complex_training_distribution_weight["0-10%"],
                                                         intensity_zone_2=complex_training_distribution_weight["11-20%"],
                                                         intensity_zone_3=complex_training_distribution_weight["21-30%"],
                                                         intensity_zone_4=complex_training_distribution_weight["31-40%"],
                                                         intensity_zone_5=complex_training_distribution_weight["41-50%"],
                                                         intensity_zone_6=complex_training_distribution_weight["51-60%"],
                                                         intensity_zone_7=complex_training_distribution_weight["61-70%"],
                                                         intensity_zone_8=complex_training_distribution_weight["71-80%"],
                                                         intensity_zone_9=complex_training_distribution_weight["81-90%"],
                                                         intensity_zone_10=complex_training_distribution_weight["91-100%"],
                                                         intensity_zone_11=complex_training_distribution_weight["101-110%"],
                                                         intensity_zone_12=complex_training_distribution_weight["111-120%"],
                                                         intensity_zone_13=complex_training_distribution_weight["121-130%"],
                                                         week=week)
                    volume3.save()
                    volume4.save()
                continue

            for key_weights in weekly_weight.keys():
                for key_one_rm in one_RM.keys():
                    if exercise == key_weights and exercise == key_one_rm:
                        for index, weight in enumerate(weekly_weight[exercise]):
                            for interval in intervals.items():

                                if weight in range((int(one_RM[key_one_rm] * interval[1][0])), (int(one_RM[key_one_rm] * interval[1][1]))):
                                    training_distribution_reps[interval[0]] += (wekly_reps[exercise][index])
                                    training_distribution_weight[interval[0]] += weight * (wekly_reps[exercise][index])
                                    training_distribution_weight['Total weight'] += weight * wekly_reps[exercise][index]
                                    training_distribution_reps["Total reps"] += (wekly_reps[exercise][index])
                                    total_training_distribution_reps[interval[0]] += (wekly_reps[exercise][index])
                                    total_training_distribution_reps["Total reps"] += (wekly_reps[exercise][index])
                                    total_training_distribution_weight['Total weight'] += weight * wekly_reps[exercise][index]
                                    total_training_distribution_weight[interval[0]] += weight * wekly_reps[exercise][index]


            volume = WeeklyTrainingVolum(mode_of_exercise=instance,
                                         wknr=wknr[0],
                                         exercise=exercise,
                                         total_intensity_zone=training_distribution_reps["Total reps"],
                                         intensity_zone_1=training_distribution_reps["0-10%"],
                                         intensity_zone_2=training_distribution_reps["11-20%"],
                                         intensity_zone_3=training_distribution_reps["21-30%"],
                                         intensity_zone_4=training_distribution_reps["31-40%"],
                                         intensity_zone_5=training_distribution_reps["41-50%"],
                                         intensity_zone_6=training_distribution_reps["51-60%"],
                                         intensity_zone_7=training_distribution_reps["61-70%"],
                                         intensity_zone_8=training_distribution_reps["71-80%"],
                                         intensity_zone_9=training_distribution_reps["81-90%"],
                                         intensity_zone_10=training_distribution_reps["91-100%"],
                                         intensity_zone_11=training_distribution_reps["101-110%"],
                                         intensity_zone_12=training_distribution_reps["111-120%"],
                                         intensity_zone_13=training_distribution_reps["121-130%"],
                                         week=week)

            volume5 = WeeklyTrainingVolumeWeight(mode_of_exercise=instance,
                                                 wknr=wknr[0],
                                                 exercise=exercise,
                                                 total_exercise_weight_volume=training_distribution_weight["Total weight"],
                                                 intensity_zone_1=training_distribution_weight["0-10%"],
                                                 intensity_zone_2=training_distribution_weight["11-20%"],
                                                 intensity_zone_3=training_distribution_weight["21-30%"],
                                                 intensity_zone_4=training_distribution_weight["31-40%"],
                                                 intensity_zone_5=training_distribution_weight["41-50%"],
                                                 intensity_zone_6=training_distribution_weight["51-60%"],
                                                 intensity_zone_7=training_distribution_weight["61-70%"],
                                                 intensity_zone_8=training_distribution_weight["71-80%"],
                                                 intensity_zone_9=training_distribution_weight["81-90%"],
                                                 intensity_zone_10=training_distribution_weight["91-100%"],
                                                 intensity_zone_11=training_distribution_weight["101-110%"],
                                                 intensity_zone_12=training_distribution_weight["111-120%"],
                                                 intensity_zone_13=training_distribution_weight["121-130%"],
                                                 week=week)
            volume.save()
            volume5.save()

        volume2 = WeeklyTrainingVolum(mode_of_exercise=instance,
                                      wknr=wknr[0],
                                      exercise='Total Reps',
                                      total_intensity_zone=total_training_distribution_reps["Total reps"],
                                      intensity_zone_1=total_training_distribution_reps["0-10%"],
                                      intensity_zone_2=total_training_distribution_reps["11-20%"],
                                      intensity_zone_3=total_training_distribution_reps["21-30%"],
                                      intensity_zone_4=total_training_distribution_reps["31-40%"],
                                      intensity_zone_5=total_training_distribution_reps["41-50%"],
                                      intensity_zone_6=total_training_distribution_reps["51-60%"],
                                      intensity_zone_7=total_training_distribution_reps["61-70%"],
                                      intensity_zone_8=total_training_distribution_reps["71-80%"],
                                      intensity_zone_9=total_training_distribution_reps["81-90%"],
                                      intensity_zone_10=total_training_distribution_reps["91-100%"],
                                      intensity_zone_11=total_training_distribution_reps["101-110%"],
                                      intensity_zone_12=total_training_distribution_reps["111-120%"],
                                      intensity_zone_13=total_training_distribution_reps["121-130%"],
                                      week=week)

        volume6 = WeeklyTrainingVolumeWeight(mode_of_exercise=instance,
                                             wknr=wknr[0],
                                             exercise='Total weight',
                                             total_exercise_weight_volume=total_training_distribution_weight["Total weight"],
                                             intensity_zone_1=total_training_distribution_weight["0-10%"],
                                             intensity_zone_2=total_training_distribution_weight["11-20%"],
                                             intensity_zone_3=total_training_distribution_weight["21-30%"],
                                             intensity_zone_4=total_training_distribution_weight["31-40%"],
                                             intensity_zone_5=total_training_distribution_weight["41-50%"],
                                             intensity_zone_6=total_training_distribution_weight["51-60%"],
                                             intensity_zone_7=total_training_distribution_weight["61-70%"],
                                             intensity_zone_8=total_training_distribution_weight["71-80%"],
                                             intensity_zone_9=total_training_distribution_weight["81-90%"],
                                             intensity_zone_10=total_training_distribution_weight["91-100%"],
                                             intensity_zone_11=total_training_distribution_weight["101-110%"],
                                             intensity_zone_12=total_training_distribution_weight["111-120%"],
                                             intensity_zone_13=total_training_distribution_weight["121-130%"],
                                             week=week)
        volume2.save()
        volume6.save()

        return total_training_distribution_reps

    wknr = list(set([dt.isocalendar()[1] for dt in find_dates(sheet_index)]))
    student_obj_ids = list(ResistanceExercise.objects.filter(student__name=name).values_list('id', flat=True))
    stud_obj = student_obj_ids[0]
    instance = ResistanceExercise.objects.get(pk=stud_obj)
    try:
        max_week = max([wk.wknr for wk in WeeklyTrainingVolum.objects.filter(mode_of_exercise=instance)])
    except ValueError:
        max_week = 0

    if wknr[0] > max_week:
        find_weekly_training_distribution(sheet_index)

@performance
def loop_over_sheets_in_diary(file_path, name, age, gender, typ):  # calls get_distribution_all_weeks() and get_distribution_all_days() to instanciates the daily and weekly classes

    weeks = 0

    if typ == 'all':

        workbook = load_workbook(filename=file_path)
        sheet_title = workbook.sheetnames

        for sheet_name in sheet_title:

            if sheet_name == 'RM oversikt' or sheet_name == 'Retningslinjer' or sheet_name == 'Treningsprogram':
                continue
            get_distribution_all_weeks(sheet_name, file_path, name, age, gender)
            get_distribution_all_days(sheet_name, file_path, name, age, gender)

            # create_overview()

        weeks = (len(workbook.sheetnames)) - 2  # change to 3

    elif typ == 'daily':

        workbook = load_workbook(filename=file_path)
        sheet_title = workbook.sheetnames

        for sheet_name in sheet_title:

            if sheet_name == 'RM oversikt' or sheet_name == 'Retningslinjer' or sheet_name == 'Treningsprogram':
                continue
            get_distribution_all_days(sheet_name, file_path, name, age, gender)

            # create_overview()

        weeks = (len(workbook.sheetnames)) - 2  # change to 3

    elif typ == 'weekly':

        workbook = load_workbook(filename=file_path)
        sheet_title = workbook.sheetnames

        for sheet_name in sheet_title:

            if sheet_name == 'RM oversikt' or sheet_name == 'Retningslinjer' or sheet_name == 'Treningsprogram':
                continue
            get_distribution_all_weeks(sheet_name, file_path, name, age, gender)

            # create_overview()

        weeks = (len(workbook.sheetnames)) - 2  # change to 3

    print(f'\n {weeks} training week(s) were investigated')


