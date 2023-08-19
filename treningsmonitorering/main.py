import sys
from PySide6.QtWidgets import QFileDialog, QWidget, QApplication, QLabel, QPushButton, QDialog, QVBoxLayout, QComboBox,\
    QMessageBox
import os
import django
os.environ.setdefault('DJANGO_SETTINGS_MODULE', 'treningsmonitorering.settings')
django.setup()
from db_interaction import validate_athlete, get_exercise_distribution, find_rm_data, retrieve_ovulation_and_menstruation_distribution
from trening_db.models import Student, TrainingRepMax
from training_data_analys import loop_over_sheets_in_diary
from openpyxl import load_workbook


def investigate_single_file(typ):
    directory_path = ''
    f = ''

    try:
        file_dialog = QFileDialog()
        file_dialog.setWindowTitle('Choose a File')
        file_dialog.setFileMode(QFileDialog.ExistingFile)
        if file_dialog.exec() == QFileDialog.Accepted:
            f = file_dialog.selectedFiles()[0]

    except Exception:
        error = QMessageBox()
        error.setWindowTitle('Feilvalg')
        error.setText('I dette vinduet må du markere en fil du ønsker å åpne.')
        error.exec()

    workbookx = load_workbook(filename=f)
    name_sheet = workbookx['RM oversikt']
    Name = name_sheet['G1'].value
    Age = name_sheet['J1'].value
    Gender = name_sheet['M1'].value
    validate_athlete(Name, Age)
    loop_over_sheets_in_diary(f, Name, Age, Gender, typ)


def investigate_multiple_files():  # Extracts the filepath to all files within a chosen folder. Calls loop_over_sheets() to instanciate.

    directory = []

    dir_path = QFileDialog.getOpenFileNames()
    for content in dir_path:
        if type(content) == list:
            for file_p in content:
                directory.append(file_p)

    for file in directory:
        workbookx = load_workbook(filename=file)
        name_sheet = workbookx['RM oversikt']
        Name = name_sheet['G1'].value
        Age = name_sheet['J1'].value
        Gender = name_sheet['M1'].value
        validate_athlete(Name, Age)
        loop_over_sheets_in_diary(file, Name, Age, Gender, 'all')


class InvestigateSingleFiles(QWidget):
    def __init__(self, parent=None):
        super(InvestigateSingleFiles, self).__init__(parent)

        self.daily_button = QPushButton('Investigate Daily Variables')
        self.weekly_button = QPushButton('Investigate Weekly Variables')

        layout = QVBoxLayout()
        layout.addWidget(self.daily_button)
        layout.addWidget(self.weekly_button)

        self.daily_button.clicked.connect(lambda: investigate_single_file('daily'))
        self.weekly_button.clicked.connect(lambda: investigate_single_file('weekly'))

        self.setLayout(layout)


class OverviewTrainingDiaries(QWidget):
    def __init__(self, parent=None):
        super(OverviewTrainingDiaries, self).__init__(parent)

        self.setWindowTitle('Analyseverktøy for Styrketreningsdagbøker')
        self.weight_button = QPushButton('Hent treningsdata (kg)')
        self.directory_button = QPushButton('Analyser treningsdata')
        self.rundown_button = QPushButton('Hent treningsdata (reps)')
        self.rm_progress_button = QPushButton('Hent RM progresjon')
        self.ovul_menstru_button = QPushButton('Hent Menstruasjons- og Eggløsningsdata')

        layout = QVBoxLayout(self)
        layout.addWidget(self.directory_button)
        layout.addWidget(self.rundown_button)
        layout.addWidget(self.weight_button)
        layout.addWidget(self.ovul_menstru_button)
        layout.addWidget((self.rm_progress_button))

        self.directory_button.clicked.connect(lambda: investigate_multiple_files())
        self.rundown_button.clicked.connect(lambda: re_analysis_dialog('dist reps'))
        self.ovul_menstru_button.clicked.connect(lambda: ovulation_menstruation_dialog())
        self.rm_progress_button.clicked.connect(lambda: re_analysis_dialog('rm'))
        self.weight_button.clicked.connect(lambda: re_analysis_dialog('dist weight'))

        def re_analysis(name=None, exercise=None, reps=None, type=None):

            student = name.currentText()
            ex = exercise.currentText()

            if type == 'dist reps':
                get_exercise_distribution(name=student, exercise=ex, type='dist reps')

            elif type == 'rm':
                rep = int(reps.currentText())
                find_rm_data(name=student, exercise=ex, reps=rep)

            elif type == 'dist weight':
                get_exercise_distribution(name=student, exercise=ex, type='dist weight')

        def fertility_analysis(name=None):

            student = name.currentText()
            retrieve_ovulation_and_menstruation_distribution(name=student)


        def re_analysis_dialog(type=None):
            analysis_box = QDialog()
            layout = QVBoxLayout(analysis_box)
            students = [stud.name for stud in Student.objects.all()]
            exercises = set([ex.exercise for ex in TrainingRepMax.objects.all()])

            name_tag = QLabel('Athletes')
            drop_down_stud = QComboBox()
            for student in students:
                drop_down_stud.addItem(student)

            exercise_tag = QLabel('Exercises')
            drop_down_ex = QComboBox()
            drop_down_ex.addItem('All exercises')
            drop_down_ex.addItem(('Total Reps'))
            for ex in exercises:
                drop_down_ex.addItem(ex)

            analyse_button = QPushButton('Summarize')

            layout.addWidget(name_tag)
            layout.addWidget(drop_down_stud)
            layout.addWidget(exercise_tag)
            layout.addWidget(drop_down_ex)

            if type == 'dist reps':
                layout.addWidget(analyse_button)
                analyse_button.clicked.connect(
                    lambda: re_analysis(name=drop_down_stud, exercise=drop_down_ex, type='dist reps'))
                analysis_box.setLayout(layout)
                analysis_box.exec()

            elif type == 'dist weight':
                layout.addWidget(analyse_button)
                analyse_button.clicked.connect(
                    lambda: re_analysis(name=drop_down_stud, exercise=drop_down_ex, type='dist weight'))
                analysis_box.setLayout(layout)
                analysis_box.exec()

            elif type == 'rm':
                re_tag = QLabel('RM')
                re_drop_down = QComboBox()
                for num in range(20):
                    number = str(num)
                    re_drop_down.addItem(number)

                layout.addWidget(re_tag)
                layout.addWidget(re_drop_down)

                layout.addWidget(analyse_button)
                analyse_button.clicked.connect(
                    lambda: re_analysis(name=drop_down_stud, exercise=drop_down_ex, reps=re_drop_down, type='rm'))
                analysis_box.setLayout(layout)
                analysis_box.exec()

       #  self.setLayout(layout)
        def ovulation_menstruation_dialog(type=None):
            analysis_box = QDialog()
            layout = QVBoxLayout(analysis_box)
            students = [stud.name for stud in Student.objects.all()]

            name_tag = QLabel('Athletes')
            drop_down_stud = QComboBox()
            for student in students:
                drop_down_stud.addItem(student)

            analyse_button = QPushButton('Summarize')

            layout.addWidget(name_tag)
            layout.addWidget(drop_down_stud)
            analyse_button.clicked.connect(lambda: fertility_analysis(name=drop_down_stud))
            layout.addWidget(analyse_button)
            analysis_box.setLayout(layout)
            analysis_box.exec()

        self.setLayout(layout)

if __name__ == '__main__':

    app = QApplication()
    window = OverviewTrainingDiaries()
    window.show()

    sys.exit(app.exec())



